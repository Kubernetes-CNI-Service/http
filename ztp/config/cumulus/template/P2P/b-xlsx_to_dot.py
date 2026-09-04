#!/usr/bin/env python3
"""
Usage: python3 b-xlsx_to_dot.py [-y] [--air-template FILE] [--os-version VERSION]
                                    [--air-link-policy FILE]

The inventory and port-mapping files are loaded automatically from the same
directory as this script. Breakout fanout is inferred globally from all P2P
link rows, including rows whose peer device is ``NA``.

Reads TAN/OOB sheets from the P2P xlsx file, resolves device ports via the
inventory / port-mapping files plus global splitter inference, and writes DOT topologies plus
an NVIDIA AIR import JSON file.
The input is always p2p.xlsx beside this script; DAY0 setup manages that link.

Column auto-detection:
  Row 1: section labels (e.g. "Source", "Dest")
  Row 2: column names   (e.g. "name", "port", "HCA/port" …)
  Fallback: columns 6/7/12/13 (0-indexed) if detection fails.

Port resolution pipeline per device+port:
  1. Match device name against inventory glob patterns (case-insensitive) → device type
  2. Direct lookup: type+port in port-mapping (e.g. "weka:P2" → "enp13s0np0")
  3. Switch pattern: group every device+base-port globally and infer the
     splitter type (1to1/1to2/1to4/1to8) from all observed suffixes
     → build pattern key (e.g. "1to4#/1/1") → match port-mapping template → substitute "#"
  4. Fallback: return original port label unchanged
"""

import csv
import fnmatch
import hashlib
import ipaddress
import json
import os
import re
import select
import sys
from copy import deepcopy

_AUTO_YES = False  # 由 -y 参数设置


def _timed_input(prompt, timeout=10, default=""):
    """Print prompt and wait up to timeout seconds for input.
    Returns default if no input arrives within the timeout.
    -y 模式下直接返回 default，不等待。
    """
    if _AUTO_YES:
        sys.stdout.write(prompt + f"{default}（-y 模式自动确认）\n")
        sys.stdout.flush()
        return default
    sys.stdout.write(prompt)
    sys.stdout.flush()
    try:
        ready, _, _ = select.select([sys.stdin], [], [], timeout)
    except (OSError, ValueError):
        print(f"\n（{timeout}s 内无输入，使用默认值：{default!r}）")
        return default
    if ready:
        try:
            line = sys.stdin.readline()
            return line.rstrip("\n") if line else default
        except EOFError:
            return default
    print(f"\n（{timeout}s 内无输入，使用默认值：{default!r}）")
    return default


try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

# ── xlsx extraction config ────────────────────────────────────────────────────

SHEET_PREFIXES       = ("TAN", "OOB")
FALLBACK_COLS        = [6, 7, 12, 13]
SECTION_SRC_KEYWORDS = ("source", "src")
SECTION_DST_KEYWORDS = ("dest", "dst")
NAME_KEYWORDS        = ("name",)
PORT_KEYWORDS        = ("port",)


def cell_str(row, idx):
    if idx >= len(row):
        return ""
    v = row[idx]
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v)


def detect_field_cols(row1, row2):
    def val(cell):
        return str(cell).strip().lower() if cell is not None else ""

    r1 = [val(c) for c in row1]
    r2 = [val(c) for c in row2]
    n  = max(len(r1), len(r2))

    src_start = next((i for i, v in enumerate(r1) if any(k in v for k in SECTION_SRC_KEYWORDS)), None)
    dst_start = next((i for i, v in enumerate(r1) if any(k in v for k in SECTION_DST_KEYWORDS)), None)

    if src_start is None or dst_start is None or src_start >= dst_start:
        return None

    def find_col(r, lo, hi, keywords):
        return next(
            (i for i in range(lo, min(hi, len(r)))
             if any(k == r[i] or k in r[i] for k in keywords)),
            None,
        )

    src_name = find_col(r2, src_start, dst_start, NAME_KEYWORDS)
    src_port = find_col(r2, src_start, dst_start, PORT_KEYWORDS)
    dst_name = find_col(r2, dst_start, n,         NAME_KEYWORDS)
    dst_port = find_col(r2, dst_start, n,         PORT_KEYWORDS)

    if None in (src_name, src_port, dst_name, dst_port):
        return None
    return [src_name, src_port, dst_name, dst_port]


def extract_xlsx(xlsx_path):
    """Return list of (src_dev, src_port, dst_dev, dst_port) tuples from xlsx."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheets = [s for s in wb.sheetnames if s.upper().startswith(SHEET_PREFIXES)]
    if not sheets:
        print("WARNING: no TAN/OOB sheets found in xlsx", file=sys.stderr)

    records = []
    for sheet_name in sheets:
        ws   = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            row1 = next(rows)
            row2 = next(rows)
        except StopIteration:
            continue

        field_cols = detect_field_cols(row1, row2)
        if field_cols is None:
            print(f"  WARNING: 无法从表头自动检测列位置，使用默认列 {FALLBACK_COLS}（sheet: {sheet_name}）",
                  file=sys.stderr)
            field_cols = FALLBACK_COLS

        count = 0
        for row in rows:
            if all(v is None for v in row):
                continue
            if len(row) > 1 and row[1] == "#N/A":
                continue
            fields = [cell_str(row, i) for i in field_cols]
            if not all(fields):
                continue
            records.append(tuple(fields))
            count += 1
        print(f"  {sheet_name:30s}  {count:4d} rows")

    wb.close()
    return records


# ── topology resolution ───────────────────────────────────────────────────────

def load_inventory(path):
    patterns   = {}
    type_order = []
    in_meta    = False
    cur_type   = None
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                in_meta = False
                continue
            if line.startswith("#"):
                continue
            if line.startswith("[[") and line.endswith("]]"):
                in_meta  = True
                cur_type = None
                continue
            if line.startswith("[") and line.endswith("]"):
                if in_meta:
                    continue
                cur_type = line[1:-1]
                if cur_type not in patterns:
                    patterns[cur_type] = []
                    type_order.append(cur_type)
                continue
            if cur_type is None or in_meta:
                continue
            pat = line.strip()
            if pat:
                patterns[cur_type].append(pat)
    return patterns, type_order


def load_port_map(path):
    direct  = {}
    switch  = {}
    cur_sec = None
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if line.startswith("[") and line.endswith("]"):
                cur_sec = line[1:-1]
                continue
            if cur_sec is None or "," not in line:
                continue
            pattern, os_name = line.split(",", 1)
            pattern, os_name = pattern.strip(), os_name.strip()
            if "#" in pattern:
                switch[f"{cur_sec}:{pattern}"] = os_name
            else:
                direct[f"{cur_sec}:{pattern}"] = os_name
    return direct, switch


def get_device_type(dev, inv_patterns, type_order):
    dev_lower = dev.lower()
    for dtype in type_order:
        for pat in inv_patterns.get(dtype, []):
            if fnmatch.fnmatch(dev_lower, pat.lower()):
                return dtype
    return "unknown"


_PLACEHOLDER_DEVICES = {"", "-", "na", "n/a", "#n/a", "none", "null", "empty"}


def _placeholder_device(name):
    """Return True for peer placeholders that must not enter DOT output."""
    return str(name or "").strip().casefold() in _PLACEHOLDER_DEVICES


def infer_splitter_profiles(records, inv_patterns, type_order,
                            port_direct, port_switch):
    """Infer breakout fanout for every device/base-port from all P2P rows.

    A breakout is valid even when only some lanes have physical peers in P2P.
    Its notation determines the hardware breakout mode; absent lanes are later
    represented as unconnected AIR interfaces and are not validation errors.
    A row whose peer device is ``NA`` still contributes its real endpoint.
    """
    grouped = {}
    unsplit = set()
    display_names = {}
    warnings = []
    switch_types = {
        key.split(":", 1)[0] for key in port_switch if ":" in key
    }

    for record in records:
        for dev, port in ((record[0], record[1]), (record[2], record[3])):
            if _placeholder_device(dev):
                continue
            dev_type = get_device_type(dev, inv_patterns, type_order)
            if dev_type not in switch_types or f"{dev_type}:{port}" in port_direct:
                continue
            parts = [part.strip() for part in str(port).split("/")]
            if len(parts) == 1:
                key = (str(dev).casefold(), parts[0])
                unsplit.add(key)
                display_names.setdefault(key, str(dev))
                continue
            base, suffix = parts[0], tuple(parts[1:])
            key = (str(dev).casefold(), base)
            grouped.setdefault(key, set()).add(suffix)
            display_names.setdefault(key, str(dev))

    profiles = {}
    for key, suffixes in sorted(
        grouped.items(),
        key=lambda item: (_natural_key(item[0][0]), _natural_key(item[0][1])),
    ):
        dev, base = display_names[key], key[1]
        if key in unsplit:
            warnings.append(
                f"{dev}:{base} 同时出现未拆分端口和拆分支路；保留原始端口名"
            )
            continue
        depths = {len(suffix) for suffix in suffixes}
        if len(depths) != 1 or next(iter(depths)) not in {1, 2}:
            pretty = ", ".join("/".join(value) for value in sorted(suffixes))
            warnings.append(
                f"{dev}:{base} 拆分编号层级冲突或不受支持：{pretty}；保留原始端口名"
            )
            continue

        depth = next(iter(depths))
        try:
            numeric = [tuple(int(part) for part in suffix) for suffix in suffixes]
        except ValueError:
            pretty = ", ".join("/".join(value) for value in sorted(suffixes))
            warnings.append(
                f"{dev}:{base} 包含非数字拆分编号：{pretty}；保留原始端口名"
            )
            continue

        if depth == 1:
            splitter_type = "1to2"
            valid = all(1 <= value[0] <= 2 for value in numeric)
        else:
            valid = all(
                1 <= value[0] <= 2 and 1 <= value[1] <= 4
                for value in numeric
            )
            splitter_type = (
                "1to8" if any(value[1] > 2 for value in numeric) else "1to4"
            )
        if not valid:
            pretty = ", ".join("/".join(value) for value in sorted(suffixes))
            warnings.append(
                f"{dev}:{base} 拆分编号超出支持范围：{pretty}；保留原始端口名"
            )
            continue

        profiles[key] = splitter_type
        dev_type = get_device_type(dev, inv_patterns, type_order)
        unsupported = [
            f"{base}/{'/'.join(suffix)}"
            for suffix in sorted(suffixes)
            if f"{dev_type}:{splitter_type}#/{'/'.join(suffix)}" not in port_switch
        ]
        if unsupported:
            warnings.append(
                f"{dev}:{base} 的 {splitter_type} 缺少端口映射："
                + ", ".join(unsupported)
            )
    return profiles, warnings


def resolve_port(dev, port, inv_patterns, type_order, port_direct, port_switch,
                 splitter_profiles):
    dev_type  = get_device_type(dev, inv_patterns, type_order)
    key       = f"{dev_type}:{port}"
    if key in port_direct:
        return port_direct[key]

    base_port = port.split("/")[0]
    splitter_type = splitter_profiles.get((str(dev).casefold(), base_port))
    if "/" in port and splitter_type is None:
        return port
    if splitter_type is None:
        splitter_type = "1to1"

    if "/" in port:
        pattern_key = f"{splitter_type}#/{port.split('/', 1)[1]}"
    else:
        pattern_key = f"{splitter_type}#"

    key = f"{dev_type}:{pattern_key}"
    if key in port_switch:
        return port_switch[key].replace("#", base_port)

    return port


def _complete_splitter_interface_inventory(
    records, splitter_profiles, inv_patterns, type_order, port_direct, port_switch
):
    """Return every hardware lane for each inferred breakout mode.

    P2P may intentionally connect only part of a breakout. The regular LLDPQ
    edge list contains only real links, while AIR JSON must still expose the
    other lanes as unconnected so full production YAML can be applied.
    """
    suffixes = {
        "1to2": [(str(lane),) for lane in range(1, 3)],
        "1to4": [
            (str(group), str(lane))
            for group in range(1, 3) for lane in range(1, 3)
        ],
        "1to8": [
            (str(group), str(lane))
            for group in range(1, 3) for lane in range(1, 5)
        ],
    }
    display_names = {}
    for record in records:
        for device, _port in ((record[0], record[1]), (record[2], record[3])):
            if not _placeholder_device(device):
                display_names.setdefault(str(device).casefold(), str(device))

    inventory = {}
    for (device_key, base), splitter_type in splitter_profiles.items():
        device = display_names.get(device_key, device_key)
        for suffix in suffixes[splitter_type]:
            raw_port = f"{base}/{'/'.join(suffix)}"
            interface = resolve_port(
                device, raw_port, inv_patterns, type_order,
                port_direct, port_switch, splitter_profiles,
            )
            inventory.setdefault(device, set()).add(interface)
    return inventory


def _configured_bond_member_inventory(csv_file):
    """Return every physical switch port referenced by production config.

    In addition to bond members, BGP, peerlink and every repeated VLAN-port
    column can require breakout lanes that have no physical P2P peer.  These
    ports must still exist in AIR before the full production YAML is applied.
    """
    inventory = {}
    if not os.path.isfile(csv_file):
        return inventory
    with open(csv_file, newline="", encoding="utf-8-sig") as stream:
        reader = csv.reader(stream)
        try:
            header = [str(value).strip().casefold() for value in next(reader)]
        except StopIteration:
            return inventory
        try:
            hostname_index = header.index("hostname")
            type_index = header.index("type")
        except ValueError as error:
            raise ValueError(
                f"{csv_file}: missing hostname/type header"
            ) from error
        port_indexes = [
            index for index, name in enumerate(header)
            if name in {"bgp_ports", "bond_ports", "peerlink_ports", "vlan_ports"}
        ]
        for row_number, row in enumerate(reader, 2):
            device = row[hostname_index].strip() if hostname_index < len(row) else ""
            device_type = (
                row[type_index].strip().casefold() if type_index < len(row) else ""
            )
            if not device or device_type == "air":
                continue
            for column_index in port_indexes:
                spec = row[column_index].strip() if column_index < len(row) else ""
                if not spec or spec.upper() == "NA":
                    continue
                for token in re.split(r"[|/]", spec):
                    token = token.strip()
                    if not token:
                        continue
                    # Schema v2 compact local-bond syntax uses one ``bond``
                    # prefix followed by ``b``-separated member numbers, for
                    # example bond49b51b53 -> swp49 + swp51 + swp53.
                    if re.fullmatch(r"bond\d+(?:b\d+)+", token, re.IGNORECASE):
                        for member in re.findall(r"\d+", token):
                            inventory.setdefault(device, set()).add(
                                f"swp{member}"
                            )
                        continue
                    # A local bond may concatenate multiple member numbers in
                    # its name (for example bond50bond52 -> swp50 + swp52).
                    if re.fullmatch(r"(?:bond\d+(?:s\d+)?){2,}", token):
                        for member in re.findall(r"bond(\d+(?:s\d+)?)", token):
                            inventory.setdefault(device, set()).add(
                                f"swp{member}"
                            )
                        continue
                    match = re.fullmatch(
                        r"(?:bond|swp)(\d+)(?:-(\d+))?"
                        r"(?:s(\d+)(?:-(\d+))?)?",
                        token,
                        re.IGNORECASE,
                    )
                    if not match:
                        # Non-physical aggregate aliases do not define switch
                        # hardware ports and are intentionally ignored.
                        if re.fullmatch(r"[A-Za-z][A-Za-z0-9_-]*", token):
                            continue
                        raise ValueError(
                            f"{csv_file}:{row_number}: unsupported port "
                            f"token {token!r}"
                        )
                    port_start, port_end, lane_start, lane_end = match.groups()
                    first_port = int(port_start)
                    last_port = int(port_end or port_start)
                    if last_port < first_port:
                        raise ValueError(
                            f"{csv_file}:{row_number}: reversed bond range {token!r}"
                        )
                    for port in range(first_port, last_port + 1):
                        if lane_start is None:
                            inventory.setdefault(device, set()).add(f"swp{port}")
                            continue
                        first_lane = int(lane_start)
                        last_lane = int(lane_end or lane_start)
                        if last_lane < first_lane:
                            raise ValueError(
                                f"{csv_file}:{row_number}: reversed bond lane "
                                f"range {token!r}"
                            )
                        for lane in range(first_lane, last_lane + 1):
                            inventory.setdefault(device, set()).add(
                                f"swp{port}s{lane}"
                            )
    return inventory


def _merge_port_inventories(*inventories):
    merged = {}
    for inventory in inventories:
        for device, interfaces in inventory.items():
            merged.setdefault(device, set()).update(interfaces)
    return merged


# ── main ──────────────────────────────────────────────────────────────────────

SCRIPT_DIR        = os.path.dirname(os.path.abspath(__file__))

DEFAULT_INV       = os.path.join(SCRIPT_DIR, "01-inventory.log")
DEFAULT_PORT_MAP  = os.path.join(SCRIPT_DIR, "02-port-mapping.log")
DEVICES_CONFIG    = os.path.abspath(
    os.path.join(SCRIPT_DIR, "..", "02-devices_config.csv")
)
LLDPQ_TEMPLATE    = os.path.join(SCRIPT_DIR, "lldpq-template.dot")
AIR_JSON_TEMPLATE = os.path.join(SCRIPT_DIR, "air-template-no-oob.json")
_AIR_DEFAULT_OS_VERSION = "5.16.4"
_AIR_OS_VERSION_RE = re.compile(
    r"^[0-9][0-9A-Za-z]{0,31}(?:[._+-][0-9A-Za-z]{1,32})*$"
)
_MAX_AIR_POLICY_BYTES = 1024 * 1024
_MAX_AIR_POLICY_REWRITES = 256


def _usage():
    return (
        f"Usage: {sys.argv[0]} [-y] [--air-template FILE] "
        "[--os-version VERSION] [--air-link-policy FILE]"
    )


def _parse_cli_args(args):
    """Parse value options without weakening unknown-option checks."""
    air_template = AIR_JSON_TEMPLATE
    os_version = None
    air_link_policy = None
    seen = set()
    index = 0
    while index < len(args):
        argument = args[index]
        if argument in {
            "--air-template", "--os-version", "--air-link-policy",
        }:
            option = argument
            index += 1
            if index >= len(args) or not args[index]:
                raise ValueError(f"{option} requires a non-empty value")
            value = args[index]
        elif argument.startswith("--air-template="):
            option = "--air-template"
            value = argument.split("=", 1)[1]
            if not value:
                raise ValueError("--air-template requires a non-empty value")
        elif argument.startswith("--os-version="):
            option = "--os-version"
            value = argument.split("=", 1)[1]
            if not value:
                raise ValueError("--os-version requires a non-empty value")
        elif argument.startswith("--air-link-policy="):
            option = "--air-link-policy"
            value = argument.split("=", 1)[1]
            if not value:
                raise ValueError(
                    "--air-link-policy requires a non-empty value"
                )
        else:
            raise ValueError(f"unknown argument: {argument}")
        if option in seen:
            raise ValueError(f"duplicate argument: {option}")
        seen.add(option)
        if option == "--air-template":
            air_template = value
        elif option == "--os-version":
            os_version = value
        else:
            air_link_policy = value
        index += 1
    return air_template, os_version, air_link_policy


def _validate_air_os_version(value):
    """Return a DOT-safe AIR Cumulus version or reject it fail closed."""
    version = str(value or "").strip()
    if len(version) > 96 or not _AIR_OS_VERSION_RE.fullmatch(version):
        raise ValueError(
            f"AIR OS version {version!r} is invalid; expected a version such as 5.18"
        )
    return version


def _missing_required_files(paths):
    """Return required inputs that are missing or are not regular files."""
    return [path for path in paths if not os.path.isfile(path)]


def _lldpq_header(template_file):
    """Return the /* ... */ comment block from the lldpq template."""
    lines, in_comment = [], False
    try:
        with open(template_file, encoding="utf-8") as f:
            for line in f:
                if line.strip().startswith("/*"):
                    in_comment = True
                if in_comment:
                    lines.append(line)
                    if line.strip().endswith("*/"):
                        break
    except OSError:
        pass
    return "".join(lines)


def _source_workbook_stem(xlsx_path):
    """Return the real workbook stem even when the fixed input is a symlink."""
    real_path = os.path.realpath(xlsx_path)
    return os.path.splitext(os.path.basename(real_path))[0]


# ── topology sort ─────────────────────────────────────────────────────────────
# Priority order: fw, oobofoob, border, oob-core, oob-spine, oob-leaf,
#                 tan-spine, tan-leaf, (everything else = servers)

def _natural_key(s):
    """Split string into (text, int, text, int, ...) for natural sort order."""
    return [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', s)]


def _dev_rank(name):
    """Return numeric sort rank for a device name.

    Order: fw → oobofoob-spine → oobofoob-leaf → border →
           oob-core → oob-spine → oob-leaf → tan-spine → tan-leaf → other
    """
    n = name.lower()
    if re.search(r'\bfw\b|^fw|-fw\d', n):             return 0
    if "oobofoob" in n and "spine" in n:               return 1
    if "oobofoob" in n:                                return 2   # oobofoob-leaf
    if "border" in n:                                  return 3
    if "oob" in n and "core" in n:                     return 4
    if "oob" in n and "spine" in n:                    return 5
    if "oob" in n and "leaf" in n:                     return 6
    if "tan" in n and "spine" in n:                    return 7
    if "tan" in n and "leaf" in n:                     return 8
    return 9   # servers / unknown


def _normalize_edge(src, sp, dst, dp):
    """Return edge with higher-priority (lower rank) device first.
    Ties broken by device name alphabetically.
    """
    ra, rb = _dev_rank(src), _dev_rank(dst)
    if ra < rb or (ra == rb and _natural_key(src) <= _natural_key(dst)):
        return src, sp, dst, dp
    return dst, dp, src, sp


def _edge_sort_key(edge):
    """Sort key for a (src, src_port, dst, dst_port) edge tuple.

    Sort by (src_rank, dst_rank) first so edges are grouped by device-type pair,
    then by device names and ports within each group.
    """
    src, sp, dst, dp = _normalize_edge(*edge)
    return (_dev_rank(src), _dev_rank(dst), _natural_key(src), _natural_key(sp), _natural_key(dst), _natural_key(dp))


_SAFE_AIR_HOSTNAME_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._-]{0,252}$")


def _validate_dot_token(value, *, context):
    """Reject workbook text that can terminate a quoted DOT identifier."""
    text = str(value)
    if not text or any(character in text for character in ('"', "\\", "\r", "\n", "\0")):
        raise ValueError(f"{context} 含不安全 DOT 字符: {text!r}")
    if any(ord(character) < 32 for character in text):
        raise ValueError(f"{context} 含控制字符: {text!r}")
    return text


def _normalized_device_name(name):
    """Return the comparison identity used for P2P device endpoints."""
    return str(name or "").strip().casefold()


def _find_duplicate_or_conflicting_links(
    records, *, allowed_self_link_indices=(),
):
    """Return indices and messages for unsafe P2P link records.

    Besides repeated links and reused ports, a record whose two endpoint
    device names normalize to the same identity is always invalid, even when
    the ports differ.  Self-link diagnostics contain only that record number
    and its two endpoints.
    """
    port_peer = {}   # (normalized device, port) -> (idx, peer device, peer port)
    link_seen = {}   # frozenset endpoint pair -> first idx
    conflict_indices = set()
    messages = []
    allowed_self_link_indices = set(allowed_self_link_indices)

    for idx, (src_dev, src_port, dst_dev, dst_port) in enumerate(records):
        src_name = _normalized_device_name(src_dev)
        dst_name = _normalized_device_name(dst_dev)

        if (not _placeholder_device(src_dev)
                and not _placeholder_device(dst_dev)
                and src_name == dst_name
                and idx not in allowed_self_link_indices):
            messages.append(
                f"  [自连接] 记录 #{idx+1}: "
                f"{str(src_dev).strip()}:{str(src_port).strip()} -- "
                f"{str(dst_dev).strip()}:{str(dst_port).strip()}"
            )
            conflict_indices.add(idx)
            continue

        link_key = frozenset({
            (src_name, src_port),
            (dst_name, dst_port),
        })

        # Case 1: exact duplicate link
        if link_key in link_seen:
            first = link_seen[link_key]
            messages.append(
                f"  [重复链路] {src_dev}:{src_port} -- {dst_dev}:{dst_port}"
                f"  (记录 #{first+1} 与 #{idx+1})"
            )
            conflict_indices.add(first)
            conflict_indices.add(idx)
        else:
            link_seen[link_key] = idx

        # Case 2: same port -> different peer
        for (dev, port), (peer_dev, peer_port) in [
            ((src_dev, src_port), (dst_dev, dst_port)),
            ((dst_dev, dst_port), (src_dev, src_port)),
        ]:
            if _placeholder_device(dev):
                continue
            ep_key = (_normalized_device_name(dev), port)
            if ep_key in port_peer:
                first_idx, first_peer_dev, first_peer_port = port_peer[ep_key]
                if (
                    _normalized_device_name(first_peer_dev), first_peer_port
                ) != (_normalized_device_name(peer_dev), peer_port):
                    messages.append(
                        f"  [端口冲突] {dev}:{port}"
                        f"  记录 #{first_idx+1}: -- "
                        f"{first_peer_dev}:{first_peer_port}"
                        f"  记录 #{idx+1}: -- {peer_dev}:{peer_port}"
                    )
                    conflict_indices.add(first_idx)
                    conflict_indices.add(idx)
            else:
                port_peer[ep_key] = (idx, peer_dev, peer_port)

    return conflict_indices, messages


def _reject_duplicate_json_keys(pairs):
    result = {}
    for key, value in pairs:
        if key in result:
            raise ValueError(f"duplicate JSON key: {key}")
        result[key] = value
    return result


def _unknown_policy_keys(value, allowed, *, context):
    unknown = sorted(set(value) - set(allowed))
    if unknown:
        raise ValueError(
            f"AIR topology policy {context} has unknown key: {unknown[0]}"
        )


def _policy_endpoint(value, *, context):
    if not isinstance(value, dict):
        raise ValueError(f"AIR topology policy {context} must be an object")
    _unknown_policy_keys(value, {"device", "port"}, context=context)
    missing = [key for key in ("device", "port") if key not in value]
    if missing:
        raise ValueError(
            f"AIR topology policy {context} is missing {missing[0]}"
        )
    if not isinstance(value["device"], str) or not isinstance(value["port"], str):
        raise ValueError(
            f"AIR topology policy {context} device and port must be strings"
        )
    device = value["device"].strip()
    port = value["port"].strip()
    if _placeholder_device(device) or not port:
        raise ValueError(
            f"AIR topology policy {context} device and port must be non-empty"
        )
    _validate_dot_token(device, context=f"AIR topology policy {context} device")
    _validate_dot_token(port, context=f"AIR topology policy {context} port")
    return {"device": device, "port": port}


def _policy_edge_signature(edge):
    return frozenset({
        (
            _normalized_device_name(edge[0]),
            str(edge[1]).strip().casefold(),
        ),
        (
            _normalized_device_name(edge[2]),
            str(edge[3]).strip().casefold(),
        ),
    })


def _policy_rule_edge(rule, field):
    endpoints = rule[field]
    return (
        endpoints[0]["device"], endpoints[0]["port"],
        endpoints[1]["device"], endpoints[1]["port"],
    )


def load_air_topology_policy(path):
    """Load a strict project-scoped policy for AIR-only topology changes."""
    policy_path = os.path.abspath(os.fspath(path))
    if not os.path.isfile(policy_path):
        raise ValueError(f"AIR topology policy file not found: {policy_path}")
    if os.path.getsize(policy_path) > _MAX_AIR_POLICY_BYTES:
        raise ValueError("AIR topology policy exceeds the 1 MiB size limit")
    try:
        with open(policy_path, encoding="utf-8") as stream:
            document = json.load(
                stream, object_pairs_hook=_reject_duplicate_json_keys,
            )
    except (OSError, UnicodeError, json.JSONDecodeError) as exc:
        raise ValueError(f"invalid AIR topology policy JSON: {exc}") from exc
    if not isinstance(document, dict):
        raise ValueError("AIR topology policy root must be an object")
    _unknown_policy_keys(
        document, {"node_allowlist", "link_rewrites"}, context="root",
    )

    raw_allowlist = document.get("node_allowlist", {})
    if not isinstance(raw_allowlist, dict):
        raise ValueError("AIR topology policy node_allowlist must be an object")
    node_allowlist = {}
    seen_allowed_nodes = set()
    for inventory_type, raw_names in raw_allowlist.items():
        if not isinstance(inventory_type, str) or not inventory_type.strip():
            raise ValueError(
                "AIR topology policy node_allowlist keys must be non-empty strings"
            )
        inventory_type = inventory_type.strip()
        if not isinstance(raw_names, list):
            raise ValueError(
                f"AIR topology policy node_allowlist.{inventory_type} must be a list"
            )
        names = []
        for index, raw_name in enumerate(raw_names, start=1):
            if not isinstance(raw_name, str) or _placeholder_device(raw_name):
                raise ValueError(
                    f"AIR topology policy node_allowlist.{inventory_type}[{index}] "
                    "must be a non-empty device name"
                )
            name = raw_name.strip()
            _validate_dot_token(
                name,
                context=(
                    f"AIR topology policy node_allowlist.{inventory_type}"
                    f"[{index}]"
                ),
            )
            identity = _normalized_device_name(name)
            if identity in seen_allowed_nodes:
                raise ValueError(
                    f"duplicate AIR topology policy allowlist device: {name}"
                )
            seen_allowed_nodes.add(identity)
            names.append(name)
        node_allowlist[inventory_type] = names

    raw_rewrites = document.get("link_rewrites", [])
    if not isinstance(raw_rewrites, list):
        raise ValueError("AIR topology policy link_rewrites must be a list")
    if len(raw_rewrites) > _MAX_AIR_POLICY_REWRITES:
        raise ValueError(
            "AIR topology policy has too many link_rewrites (maximum 256)"
        )
    link_rewrites = []
    seen_matches = set()
    for rule_number, raw_rule in enumerate(raw_rewrites, start=1):
        context = f"link_rewrites rule #{rule_number}"
        if not isinstance(raw_rule, dict):
            raise ValueError(f"AIR topology policy {context} must be an object")
        _unknown_policy_keys(
            raw_rule, {"scope", "match", "replacement"}, context=context,
        )
        missing = [
            key for key in ("scope", "match", "replacement")
            if key not in raw_rule
        ]
        if missing:
            raise ValueError(
                f"AIR topology policy {context} is missing {missing[0]}"
            )
        if raw_rule["scope"] != "air":
            raise ValueError(
                f"AIR topology policy {context} scope must be 'air'"
            )
        rule = {"scope": "air"}
        for field in ("match", "replacement"):
            raw_endpoints = raw_rule[field]
            if not isinstance(raw_endpoints, list) or len(raw_endpoints) != 2:
                raise ValueError(
                    f"AIR topology policy {context}.{field} must have 2 endpoints"
                )
            rule[field] = [
                _policy_endpoint(
                    endpoint,
                    context=f"{context}.{field}[{endpoint_index}]",
                )
                for endpoint_index, endpoint in enumerate(raw_endpoints, start=1)
            ]
        match_edge = _policy_rule_edge(rule, "match")
        match_signature = _policy_edge_signature(match_edge)
        if len(match_signature) != 2:
            raise ValueError(
                f"AIR topology policy {context}.match repeats one endpoint"
            )
        if match_signature in seen_matches:
            raise ValueError(
                f"duplicate AIR topology policy link match in {context}"
            )
        seen_matches.add(match_signature)

        replacement_edge = _policy_rule_edge(rule, "replacement")
        if (
            _normalized_device_name(replacement_edge[0])
            == _normalized_device_name(replacement_edge[2])
        ):
            raise ValueError(
                f"AIR topology policy {context} replacement is a self-link"
            )
        link_rewrites.append(rule)

    return {
        "node_allowlist": node_allowlist,
        "link_rewrites": link_rewrites,
    }


def _air_policy_rewrite_matches(policy, edges):
    matches = []
    used_indices = set()
    edge_signatures = [_policy_edge_signature(edge) for edge in edges]
    for rule_number, rule in enumerate(
        policy.get("link_rewrites", []), start=1,
    ):
        signature = _policy_edge_signature(_policy_rule_edge(rule, "match"))
        indices = [
            index for index, edge_signature in enumerate(edge_signatures)
            if edge_signature == signature
        ]
        if len(indices) != 1:
            raise ValueError(
                f"AIR topology policy link_rewrites rule #{rule_number} "
                f"matched {len(indices)} links; expected exactly 1"
            )
        index = indices[0]
        if index in used_indices:
            raise ValueError(
                f"AIR topology policy link_rewrites rule #{rule_number} "
                "matches an edge already rewritten by another rule"
            )
        used_indices.add(index)
        matches.append((rule, index))
    return matches


def _rewritten_air_edges(edges, matches):
    replacements = {
        index: _policy_rule_edge(rule, "replacement")
        for rule, index in matches
    }
    rewritten = [replacements.get(index, edge) for index, edge in enumerate(edges)]
    conflicts, messages = _find_duplicate_or_conflicting_links(rewritten)
    if conflicts:
        detail = messages[0].strip() if messages else "unknown conflict"
        raise ValueError(
            f"AIR topology policy produces an unsafe link topology: {detail}"
        )
    return rewritten


def _validate_air_topology_policy(policy, edges, inv_patterns, type_order):
    """Bind every explicit policy selector to exactly one current topology."""
    if not policy:
        return set()
    nodes = {}
    for src, _src_port, dst, _dst_port in edges:
        for device in (src, dst):
            nodes.setdefault(_normalized_device_name(device), str(device).strip())

    for inventory_type, names in policy.get("node_allowlist", {}).items():
        if inventory_type not in inv_patterns:
            raise ValueError(
                "AIR topology policy node_allowlist references unknown inventory "
                f"type: {inventory_type}"
            )
        for name in names:
            identity = _normalized_device_name(name)
            if identity not in nodes:
                raise ValueError(
                    "AIR topology policy node_allowlist device matched 0 nodes: "
                    f"{name}"
                )
            actual_type = get_device_type(
                nodes[identity], inv_patterns, type_order,
            )
            if actual_type != inventory_type:
                raise ValueError(
                    f"AIR topology policy allowlist device {name} has inventory "
                    f"type {actual_type}, expected {inventory_type}"
                )

    matches = _air_policy_rewrite_matches(policy, edges)
    for rule_number, (rule, _index) in enumerate(matches, start=1):
        for endpoint in rule["replacement"]:
            if _normalized_device_name(endpoint["device"]) not in nodes:
                raise ValueError(
                    f"AIR topology policy link_rewrites rule #{rule_number} "
                    "replacement device matched 0 nodes: "
                    f"{endpoint['device']}"
                )
    _rewritten_air_edges(edges, matches)
    return {index for _rule, index in matches}


def _apply_air_topology_policy(edges, policy):
    if not policy or not policy.get("link_rewrites"):
        return list(edges)
    return _rewritten_air_edges(
        edges, _air_policy_rewrite_matches(policy, edges),
    )


def _air_policy_allows_node(name, inv_patterns, type_order, policy):
    if not policy:
        return True
    inventory_type = get_device_type(name, inv_patterns, type_order)
    allowlist = policy.get("node_allowlist", {})
    if inventory_type not in allowlist:
        return True
    allowed = {
        _normalized_device_name(device)
        for device in allowlist[inventory_type]
    }
    return _normalized_device_name(name) in allowed


def main():
    global _AUTO_YES
    args = sys.argv[1:]
    if "-y" in args:
        _AUTO_YES = True
        args = [a for a in args if a != "-y"]

    if args in (["-h"], ["--help"]):
        print(_usage())
        return

    try:
        (
            air_json_template,
            explicit_os_version,
            air_link_policy_file,
        ) = _parse_cli_args(args)
        if explicit_os_version is not None:
            explicit_os_version = _validate_air_os_version(explicit_os_version)
    except ValueError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        print(_usage(), file=sys.stderr)
        sys.exit(1)
    if not os.path.isabs(air_json_template):
        air_json_template = os.path.join(SCRIPT_DIR, air_json_template)
    air_json_template = os.path.abspath(air_json_template)
    if air_link_policy_file is not None:
        if not os.path.isabs(air_link_policy_file):
            air_link_policy_file = os.path.join(
                SCRIPT_DIR, air_link_policy_file,
            )
        air_link_policy_file = os.path.abspath(air_link_policy_file)

    xlsx_path = os.path.join(SCRIPT_DIR, "p2p.xlsx")
    print("使用固定输入：p2p.xlsx")

    inv_file, port_map_file = DEFAULT_INV, DEFAULT_PORT_MAP

    required_inputs = [
        xlsx_path,
        inv_file,
        port_map_file,
        air_json_template,
        DEVICES_CONFIG,
    ]
    if air_link_policy_file is not None:
        required_inputs.append(air_link_policy_file)
    missing_inputs = _missing_required_files(required_inputs)
    if missing_inputs:
        for path in missing_inputs:
            print(f"ERROR: File not found: {path}", file=sys.stderr)
        print(
            "ERROR: 02-devices_config.csv is required to preserve configured "
            "breakout ports that have no physical P2P peer.",
            file=sys.stderr,
        )
        sys.exit(1)

    print(f"使用 AIR JSON 模板：{air_json_template}")
    air_topology_policy = None
    if air_link_policy_file is not None:
        try:
            air_topology_policy = load_air_topology_policy(
                air_link_policy_file
            )
        except ValueError as exc:
            print(f"[ERROR] {exc}", file=sys.stderr)
            sys.exit(1)
        print(f"使用 AIR 拓扑策略：{air_link_policy_file}")

    out_dir = os.path.join(SCRIPT_DIR, "output-p2p")
    os.makedirs(out_dir, exist_ok=True)
    base        = _source_workbook_stem(xlsx_path)
    output_file = os.path.join(out_dir, base + "-lldpq.dot")

    # ── Extract p2p records from xlsx ────────────────────────────────────────
    print(f"Reading: {xlsx_path}")
    records = extract_xlsx(xlsx_path)
    print(f"\nTotal: {len(records)} records")

    if not records:
        print("No records found; nothing to generate.")
        sys.exit(0)

    # ── Load resolution data ─────────────────────────────────────────────────
    inv_patterns, type_order = load_inventory(inv_file)
    port_direct, port_switch = load_port_map(port_map_file)
    try:
        for lineno, record in enumerate(records, start=1):
            for label, value in zip(
                ("source device", "source port", "destination device", "destination port"),
                record,
            ):
                _validate_dot_token(value, context=f"P2P 记录 {lineno} {label}")
            for device in (record[0], record[2]):
                if (_is_eth_sw(device, inv_patterns, type_order)
                        and not _SAFE_AIR_HOSTNAME_RE.fullmatch(_air_hostname(device))):
                    raise ValueError(
                        f"P2P 记录 {lineno} 交换机名无法作为 AIR/ZTP hostname: "
                        f"{device!r}"
                    )
    except ValueError as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        sys.exit(1)
    splitter_profiles, splitter_warnings = infer_splitter_profiles(
        records, inv_patterns, type_order, port_direct, port_switch
    )
    print(f"推断出 {len(splitter_profiles)} 组拆分端口配置")
    if splitter_warnings:
        print(
            f"[WARNING] 拆分端口检查发现 {len(splitter_warnings)} 个问题：",
            file=sys.stderr,
        )
        for message in splitter_warnings:
            print(f"  {message}", file=sys.stderr)

    # ── Check for devices not in inventory ───────────────────────────────────
    unknown_devs = {}
    for lineno, (src_dev, _sp, dst_dev, _dp) in enumerate(records, 1):
        for dev in (src_dev, dst_dev):
            if _placeholder_device(dev):
                continue
            if get_device_type(dev, inv_patterns, type_order) == "unknown":
                if dev not in unknown_devs:
                    unknown_devs[dev] = lineno

    if unknown_devs:
        print("\nWARNING: 以下设备在 inventory 中未定义（将以原始端口名写入 dot 文件）：")
        for dev, lineno in sorted(unknown_devs.items()):
            print(f"  {dev}  (首次出现于第 {lineno} 行)")
        try:
            answer = _timed_input("\n是否继续生成？[y/N] ", default="y").strip().lower()
        except EOFError:
            answer = "n"
        if answer not in ("y", "yes"):
            print("已中止。")
            sys.exit(1)

    # AIR policy selectors use resolved OS ports (for example raw ``32`` is
    # matched as ``swp32``) while diagnostics keep the original record index.
    policy_edges = []
    policy_record_indices = []
    for record_index, (src_dev, src_port, dst_dev, dst_port) in enumerate(records):
        if _placeholder_device(src_dev) or _placeholder_device(dst_dev):
            continue
        policy_edges.append((
            src_dev,
            resolve_port(
                src_dev, src_port, inv_patterns, type_order,
                port_direct, port_switch, splitter_profiles,
            ),
            dst_dev,
            resolve_port(
                dst_dev, dst_port, inv_patterns, type_order,
                port_direct, port_switch, splitter_profiles,
            ),
        ))
        policy_record_indices.append(record_index)
    try:
        policy_match_indices = _validate_air_topology_policy(
            air_topology_policy, policy_edges, inv_patterns, type_order,
        )
    except ValueError as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        sys.exit(1)
    allowed_self_link_indices = {
        policy_record_indices[index] for index in policy_match_indices
    }

    # ── Detect duplicate / conflicting port usage ─────────────────────────────
    # Case 0: same normalized device on both endpoints, unless AIR rewrites it
    # Case 1: exact same link appears twice (A:p -- B:q and again A:p -- B:q)
    # Case 2: same port connected to two different peers (A:p -- B:q and A:p -- C:r)
    # Every case fails closed before generation; no record is silently omitted.
    dup_indices, dup_messages = _find_duplicate_or_conflicting_links(
        records, allowed_self_link_indices=allowed_self_link_indices,
    )

    if dup_indices:
        print(
            f"[ERROR] P2P 存在 {len(dup_indices)} 条自连接、重复或端口冲突记录；"
            "不会排除后继续生成，请先修正 XLSX 或显式 AIR 策略：",
            file=sys.stderr,
        )
        for msg in dup_messages:
            print(msg, file=sys.stderr)
        sys.exit(1)

    # ── Generate dot file ─────────────────────────────────────────────────────
    resolved = []
    ztp_bmc_endpoints = set()
    for src_dev, src_port, dst_dev, dst_port in records:
        if _placeholder_device(src_dev) or _placeholder_device(dst_dev):
            continue
        src_os = resolve_port(src_dev, src_port, inv_patterns, type_order,
                              port_direct, port_switch, splitter_profiles)
        dst_os = resolve_port(dst_dev, dst_port, inv_patterns, type_order,
                              port_direct, port_switch, splitter_profiles)
        if _air_is_ztp_server(src_dev) and src_port.strip().casefold().startswith("bmc"):
            ztp_bmc_endpoints.add((src_dev.casefold(), src_os.casefold()))
        if _air_is_ztp_server(dst_dev) and dst_port.strip().casefold().startswith("bmc"):
            ztp_bmc_endpoints.add((dst_dev.casefold(), dst_os.casefold()))
        resolved.append((src_dev, src_os, dst_dev, dst_os))

    resolved.sort(key=_edge_sort_key)
    lines = [f'"{s}":"{sp}" -- "{d}":"{dp}"'
             for s, sp, d, dp in (_normalize_edge(*e) for e in resolved)]

    network_name = base
    header       = _lldpq_header(LLDPQ_TEMPLATE)

    with open(output_file, "w", encoding="utf-8") as f:
        if header:
            f.write(header)
            f.write("\n")
        f.write(f'graph "{network_name}" {{\n\n')
        f.write("\n".join(lines))
        f.write("\n\n}\n")

    print(f"\nGenerated: {output_file}  ({len(lines)} edges)")

    inferred_port_inventory = _complete_splitter_interface_inventory(
        records, splitter_profiles, inv_patterns, type_order,
        port_direct, port_switch,
    )
    configured_port_inventory = _configured_bond_member_inventory(
        DEVICES_CONFIG
    )
    required_unconnected_ports = _merge_port_inventories(
        inferred_port_inventory, configured_port_inventory
    )

    # ── Generate AIR dot file ─────────────────────────────────────────────────
    air_file = os.path.join(out_dir, base + "-air.dot")
    air_json_file = os.path.join(out_dir, base + "-air.json")
    if explicit_os_version is not None:
        os_ver = explicit_os_version
        print(f"使用显式 AIR Cumulus OS 版本：{os_ver}")
    else:
        try:
            os_ver = _timed_input(
                "\n请输入 Cumulus OS 版本（用于 AIR simulation dot 文件中各节点的 os= 字段，"
                f"直接回车默认 {_AIR_DEFAULT_OS_VERSION}）：",
                default="",
            ).strip()
        except EOFError:
            os_ver = ""
        try:
            os_ver = _validate_air_os_version(
                os_ver or _AIR_DEFAULT_OS_VERSION
            )
        except ValueError as exc:
            print(f"[ERROR] {exc}", file=sys.stderr)
            sys.exit(1)
    generate_air_dot(
        output_file,
        air_file,
        inv_patterns,
        type_order,
        os_version=os_ver,
        template_file=air_json_template,
        ztp_bmc_endpoints=ztp_bmc_endpoints,
        required_unconnected_ports=required_unconnected_ports,
        air_topology_policy=air_topology_policy,
    )
    generate_air_json(
        air_file,
        air_json_file,
        air_json_template,
        lldpq_file=output_file,
        air_topology_policy=air_topology_policy,
    )

    print(f"DOT and AIR JSON files are ready in: {out_dir}")


# ── AIR dot generation ────────────────────────────────────────────────────────

_DOT_EDGE_RE = re.compile(r'"([^"]+)":"([^"]+)"\s+--\s+"([^"]+)":"([^"]+)"')
_AIR_NODE_RE = re.compile(r'^\s*"([^"]+)"\s*\[(.*?)\]\s*$', re.MULTILINE)
_AIR_ATTR_RE = re.compile(
    r'([A-Za-z_][A-Za-z0-9_]*)\s*=\s*(?:"((?:\\.|[^"])*)"|([^\s\]]+))'
)


def _is_eth_sw(name, inv_patterns, type_order):
    """True if device is a network switch.

    An explicit PDU inventory match vetoes name-based switch heuristics.  This
    matters for names such as ``example-oob-corepod-pdu01`` which otherwise look
    like an OOB core.  Other legacy names still fall back to ``_dev_rank`` so
    firewalls continue to be represented in AIR.
    """
    device_type = get_device_type(name, inv_patterns, type_order)
    if device_type == "Eth-SW":
        return True
    if device_type == "PDU":
        return False
    return _dev_rank(name) < 9   # fw / border / spine / leaf / core / oobofoob


# Model lookup: ordered list of (regex_pattern, model_string).
# First match wins.
_AIR_MEMORY_MB = 4096
_AIR_CPUS = 4
_AIR_HOSTNAME_PREFIX = "AIR-"
_AIR_MODEL_RULES = [
    (re.compile(r'\bfw\b|^fw|-fw\d',               re.I), "SN2201"),
    (re.compile(r'tor',                              re.I), "SN2201"),
    (re.compile(r'tan.*spine',                      re.I), "SN5610"),
    (re.compile(r'tan.*leaf',                       re.I), "SN5610"),
    (re.compile(r'border',                           re.I), "SN4700"),
    (re.compile(r'oob.*core',                       re.I), "SN5610"),
    (re.compile(r'oob.*spine',                      re.I), "SN4700"),
    (re.compile(r'oob.*leaf',                       re.I), "SN2201"),
    (re.compile(r'oobofoob.*spine',                 re.I), "SN4700"),
    (re.compile(r'oobofoob',                        re.I), "SN2201"),
]
_AIR_MODEL_DEFAULT = "SN5610"


def _air_resources(model):
    """Return the fixed AIR size contract for a generated platform model."""
    if str(model).strip().casefold() == "sn2201":
        return 2048, 2
    return _AIR_MEMORY_MB, _AIR_CPUS


def _air_hostname(name):
    """Add the AIR namespace prefix once to a generated AIR node name."""
    value = str(name)
    return value if value.casefold().startswith(
        _AIR_HOSTNAME_PREFIX.casefold()
    ) else _AIR_HOSTNAME_PREFIX + value


def _air_model(name):
    for pattern, model in _AIR_MODEL_RULES:
        if pattern.search(name):
            return model
    return _AIR_MODEL_DEFAULT


def _air_management_port(interface):
    """Return True for management-only endpoint names used in P2P data."""
    value = str(interface).strip().casefold()
    return value == "eth0" or value.startswith("mgmt") or value.startswith("bmc")


def _air_is_ztp_server(name):
    """Recognise the real ZTP server supplied by the P2P topology."""
    return "ztp-server" in str(name).casefold()


def _air_drop_server_port(name, interface):
    """Apply management-link filtering, with the ZTP server BMC exception."""
    value = str(interface).strip().casefold()
    if _air_is_ztp_server(name):
        return value.startswith("bmc")
    return _air_management_port(interface)


def _air_ztp_server_interfaces(ztp_links):
    """Preserve real ethN server ports and map only non-AIR labels.

    The oob-mgmt-server image reserves eth0 for management.  Physical P2P
    labels such as BF-P1/BF-P2 are not Linux NIC names, so they receive the
    lowest unused ethN value after all explicit eth1+ interfaces are reserved.
    """
    explicit = set()
    for _dev, _sw_port, _server, server_port in ztp_links:
        match = re.fullmatch(r"eth([1-9]\d*)", str(server_port).strip(), re.IGNORECASE)
        if match:
            explicit.add(int(match.group(1)))

    used = set(explicit)
    next_index = 1
    result = []
    for link in ztp_links:
        server_port = str(link[3]).strip()
        match = re.fullmatch(r"eth([1-9]\d*)", server_port, re.IGNORECASE)
        if match:
            interface = f"eth{int(match.group(1))}"
        else:
            while next_index in used:
                next_index += 1
            interface = f"eth{next_index}"
            used.add(next_index)
            next_index += 1
        result.append((link, interface))
    return result


def _air_template_role(name):
    """Map a generated node name to a reference node in the AIR template."""
    lowered = str(name).casefold()
    if lowered.startswith(_AIR_HOSTNAME_PREFIX.casefold()):
        lowered = lowered[len(_AIR_HOSTNAME_PREFIX):]
    if lowered in {"host", "ztp-server"}:
        return "ztp-server"
    if "oobofoob" in lowered:
        # OOBofOOB switches use the same AIR VM shape, Cumulus image,
        # hardware label and interface inventory as regular OOB leaves.
        return "OOB-Leaf"
    if re.search(r'\bfw\b|^fw|-fw\d', lowered):
        return "FW"
    if "border" in lowered:
        return "Border01"
    if "tan" in lowered and "spine" in lowered:
        return "TAN-Spine"
    if "tan" in lowered and "leaf" in lowered:
        return "TAN-Leaf"
    if "oob" in lowered and "core" in lowered:
        return "OOB-Core"
    if "oob" in lowered and "spine" in lowered:
        return "OOB-Spine"
    if "oob" in lowered and ("leaf" in lowered or "tor" in lowered):
        return "OOB-Leaf"
    return "TAN-Leaf"


def _air_template_node(template_nodes, node_name, explicit_name=None):
    """Select the closest node prototype while supporting both AIR templates."""
    if not isinstance(template_nodes, dict) or not template_nodes:
        raise ValueError("AIR JSON template content.nodes must be a non-empty object")

    casefold_names = {str(key).casefold(): key for key in template_nodes}
    # OOBofOOB always follows OOB-Leaf. This intentionally overrides an old
    # DOT's explicit ``template_node=OOBofOOB-Leaf`` so existing topology
    # artifacts also adopt the current generation rule.
    requested = (
        "OOB-Leaf"
        if "oobofoob" in str(node_name).casefold()
        else explicit_name or _air_template_role(node_name)
    )
    key = casefold_names.get(str(requested).casefold())
    if key is not None:
        return key, deepcopy(template_nodes[key])

    # The legacy OOB template uses generic border/leaf/spine/server names.
    fallback_fragments = {
        "ztp-server": ("server",),
        "oobofoob-leaf": ("leaf",),
        "fw": ("fw",),
        "border01": ("border",),
        "tan-spine": ("spine",),
        "tan-leaf": ("leaf",),
        "oob-spine": ("spine",),
        "oob-leaf": ("leaf",),
    }
    for fragment in fallback_fragments.get(str(requested).casefold(), ()):
        for candidate, node in template_nodes.items():
            if fragment in str(candidate).casefold():
                return candidate, deepcopy(node)

    for candidate, node in template_nodes.items():
        if str(node.get("os", "")).startswith("cumulus-vx-"):
            return candidate, deepcopy(node)
    candidate = next(iter(template_nodes))
    return candidate, deepcopy(template_nodes[candidate])


def generate_air_dot(
    lldpq_file,
    air_file,
    inv_patterns,
    type_order,
    os_version=_AIR_DEFAULT_OS_VERSION,
    template_file=AIR_JSON_TEMPLATE,
    ztp_bmc_endpoints=None,
    required_unconnected_ports=None,
    air_topology_policy=None,
):
    """Create an AIR-format dot file from a lldpq dot file.

    Rules:
    - Keep all links where BOTH endpoints are Eth-SW (border/spine/leaf/core/fw),
      including eth0/mgmt/bmc links.
    - Remove regular non-network devices from AIR, but preserve the exact
      switch-side interface from each removed link as an unconnected port.
      This is required for breakout names such as swp11s0; replacing them with
      a model-level parent such as swp11 changes the VX interface inventory.
    - For the real ZTP server, drop only BMC and keep every other link.  The
      switch side of a dropped BMC link is likewise preserved as unconnected.
    - Build ``ztp-server`` only from a real P2P device whose name contains
      ``ztp-server``; retain all of its non-management links.
    - Prefix every AIR node hostname and every link endpoint with ``AIR-``;
      the source lldpq dot file keeps its original hostnames.
    - Output the full graphviz graph format with node declarations.
    """
    with open(template_file, encoding="utf-8") as stream:
        air_template = json.load(stream)
    template_content = air_template.get("content", {})
    template_nodes = template_content.get("nodes", {})
    assign_management_ips = isinstance(template_content.get("oob"), dict)
    ztp_bmc_endpoints = {
        (str(node).casefold(), str(port).casefold())
        for node, port in (ztp_bmc_endpoints or set())
    }

    edges = []
    with open(lldpq_file, encoding="utf-8") as f:
        for line in f:
            m = _DOT_EDGE_RE.match(line.strip())
            if m:
                edges.append((m.group(1), m.group(2), m.group(3), m.group(4)))
    _validate_air_topology_policy(
        air_topology_policy, edges, inv_patterns, type_order,
    )
    edges = _apply_air_topology_policy(edges, air_topology_policy)

    def is_net(name):
        return (
            _is_eth_sw(name, inv_patterns, type_order)
            and _air_policy_allows_node(
                name, inv_patterns, type_order, air_topology_policy,
            )
        )

    # Collect Eth-SW devices, sorted by type rank then name
    net_devs_seen = set()
    for src, _sp, dst, _dp in edges:
        for dev in (src, dst):
            if is_net(dev):
                net_devs_seen.add(dev)
    net_devs_ordered = sorted(net_devs_seen,
                              key=lambda d: (_dev_rank(d), d.lower()))

    # Classify links and retain only the real ZTP server among non-network nodes.
    sw_sw_links = []
    ztp_links = []
    ztp_source_names = set()
    preserved_switch_ports = {
        str(device).casefold(): set(interfaces)
        for device, interfaces in (required_unconnected_ports or {}).items()
    }

    def preserve_switch_port(device, interface):
        preserved_switch_ports.setdefault(str(device).casefold(), set()).add(interface)

    def consume_switch_port(device, interface):
        preserved_switch_ports.setdefault(str(device).casefold(), set()).discard(
            interface
        )

    for src, sp, dst, dp in edges:
        src_net = is_net(src)
        dst_net = is_net(dst)
        if src_net and dst_net:
            sw_sw_links.append((src, sp, dst, dp))
            consume_switch_port(src, sp)
            consume_switch_port(dst, dp)
            continue
        if src_net == dst_net:
            continue

        if src_net:
            net_device, net_port = src, sp
            peer_device, peer_port = dst, dp
        else:
            net_device, net_port = dst, dp
            peer_device, peer_port = src, sp

        peer_endpoint = (peer_device.casefold(), peer_port.casefold())
        if _air_is_ztp_server(peer_device):
            if (peer_endpoint in ztp_bmc_endpoints or
                    _air_drop_server_port(peer_device, peer_port)):
                preserve_switch_port(net_device, net_port)
                continue
            ztp_source_names.add(peer_device)
            ztp_links.append(
                (net_device, net_port, peer_device, peer_port)
            )
            consume_switch_port(net_device, net_port)
            continue

        # Ordinary servers are deliberately absent from AIR, but their switch
        # ports still define the simulated switch's exact interface inventory.
        preserve_switch_port(net_device, net_port)

    sw_sw_links.sort(key=_edge_sort_key)
    ztp_links.sort(key=lambda edge: (
        _dev_rank(edge[0]), _natural_key(edge[0]), _natural_key(edge[1]),
        _natural_key(edge[3]),
    ))
    if len({name.casefold() for name in ztp_source_names}) > 1:
        raise ValueError(
            "AIR topology contains more than one ZTP server: "
            + ", ".join(sorted(ztp_source_names, key=_natural_key))
        )

    # Build output lines
    out = ["graph network {", ""]

    # Node declarations — Eth-SW devices
    mgmt_base = 11
    for dev in net_devs_ordered:
        template_name, prototype = _air_template_node(template_nodes, dev)
        labels = prototype.get("labels") or {}
        model = labels.get("model") or _air_model(dev)
        node_os = str(prototype.get("os", f"cumulus-vx-{os_version}"))
        if node_os.startswith("cumulus-vx-"):
            node_os = f"cumulus-vx-{os_version}"
        # Template prototypes describe platform shape, but sizing follows the
        # generated model contract: compact SN2201, 4096/4 for other switches.
        memory, cpus = _air_resources(model)
        oob = str(bool(prototype.get("oob", False))).lower()
        air_dev = _air_hostname(dev)
        mgmt_attr = ""
        if assign_management_ips:
            mgmt_attr = f' mgmt_ip="192.168.200.{mgmt_base}"'
            mgmt_base += 1
        preserved_attr = ""
        preserved = sorted(
            preserved_switch_ports.get(dev.casefold(), set()), key=_natural_key
        )
        if preserved:
            preserved_attr = (
                ' preserved_unconnected_ports="' + ','.join(preserved) + '"'
            )
        out.append(
            f'"{air_dev}" [ memory="{memory}" model="{model}" '
            f'os="{node_os}" cpus="{cpus}" oob="{oob}" '
            f'template_node="{template_name}"{mgmt_attr}{preserved_attr} ]'
        )

    if ztp_links:
        out.append("")
        ztp_template_name, ztp_prototype = _air_template_node(
            template_nodes, "ztp-server", explicit_name="ztp-server"
        )
        ztp_labels = ztp_prototype.get("labels") or {}
        ztp_model = ztp_labels.get("model") or "SN2201"
        ztp_os = str(ztp_prototype.get("os", f"cumulus-vx-{os_version}"))
        if ztp_os.startswith("cumulus-vx-"):
            ztp_os = f"cumulus-vx-{os_version}"
        ztp_memory, ztp_cpus = _air_resources(ztp_model)
        ztp_oob = str(bool(ztp_prototype.get("oob", False))).lower()
        ztp_mgmt_attr = ""
        if assign_management_ips:
            ztp_mgmt_attr = f' mgmt_ip="192.168.200.{mgmt_base}"'
        out.append(
            f'"ztp-server" [ memory="{ztp_memory}" model="{ztp_model}" '
            f'os="{ztp_os}" cpus="{ztp_cpus}" oob="{ztp_oob}" '
            f'template_node="{ztp_template_name}"{ztp_mgmt_attr} ]'
        )
    out.append("")

    # Switch-switch links
    for edge in sw_sw_links:
        src, sp, dst, dp = _normalize_edge(*edge)
        out.append(
            f'"{_air_hostname(src)}":"{sp}" -- '
            f'"{_air_hostname(dst)}":"{dp}"'
        )

    out.append("")

    # Preserve explicit eth1+ interfaces from the source P2P. Labels such as
    # BF-P1/BF-P2 are not AIR/Linux NIC names and use free ethN interfaces.
    for (dev, sw_port, _server, _server_port), server_interface in \
            _air_ztp_server_interfaces(ztp_links):
        out.append(
            f'"{_air_hostname(dev)}":"{sw_port}" -- '
            f'"ztp-server":"{server_interface}"'
        )

    out.append("")
    out.append("}")

    with open(air_file, "w", encoding="utf-8") as f:
        f.write("\n".join(out) + "\n")

    print(f"Generated: {air_file}  "
          f"({len(sw_sw_links)} sw-sw links, {len(ztp_links)} ztp-server links)")


def _parse_air_dot(air_file):
    """Return ordered AIR DOT nodes and links."""
    with open(air_file, encoding="utf-8") as stream:
        source = stream.read()

    nodes = []
    seen_nodes = set()
    for match in _AIR_NODE_RE.finditer(source):
        name, raw_attrs = match.groups()
        if name in seen_nodes:
            raise ValueError(f"duplicate AIR node in {air_file}: {name}")
        attrs = {}
        for attr_match in _AIR_ATTR_RE.finditer(raw_attrs):
            key, quoted, unquoted = attr_match.groups()
            value = quoted if quoted is not None else unquoted
            attrs[key] = value.replace(r'\"', '"').replace(r'\\', '\\')
        nodes.append((name, attrs))
        seen_nodes.add(name)

    links = [match.groups() for match in _DOT_EDGE_RE.finditer(source)]
    if not nodes:
        raise ValueError(f"no AIR nodes found in {air_file}")
    if not links:
        raise ValueError(f"no AIR links found in {air_file}")
    for left_node, _left_port, right_node, _right_port in links:
        for node in (left_node, right_node):
            if node not in seen_nodes:
                raise ValueError(f"AIR link references undefined node {node!r}")
    return nodes, links


def _air_json_links_from_lldpq(
    lldpq_file, nodes, air_links, *, air_topology_policy=None,
):
    """Build JSON links from the full LLDPQ graph and an AIR node allowlist.

    AIR DOT owns node selection and simulation attributes.  LLDPQ DOT owns the
    physical links and exact port names.  A link whose peer was removed from
    AIR becomes an unconnected endpoint on the retained node.

    The ZTP server is the sole special case: generate_air_dot() maps its source
    P1/P2-style names to AIR eth1..ethN.  Reuse that mapping, and treat a ZTP
    link deliberately omitted from AIR DOT (notably BMC) as an unconnected
    switch endpoint instead of reintroducing the filtered server port.
    """
    selected = {}
    ztp_node = None
    for name, _attrs in nodes:
        if _air_is_ztp_server(name):
            ztp_node = name
            continue
        source_name = str(name)
        if source_name.casefold().startswith(_AIR_HOSTNAME_PREFIX.casefold()):
            source_name = source_name[len(_AIR_HOSTNAME_PREFIX):]
        key = source_name.casefold()
        if key in selected:
            raise ValueError(f"AIR node allowlist is ambiguous: {source_name}")
        selected[key] = name

    # Keyed by the non-ZTP endpoint, because the LLDPQ-side ZTP interface name
    # is intentionally replaced while producing AIR DOT.
    ztp_interfaces = {}
    if ztp_node is not None:
        for left_node, left_port, right_node, right_port in air_links:
            if left_node == ztp_node and right_node != ztp_node:
                ztp_interfaces[(right_node.casefold(), right_port.casefold())] = left_port
            elif right_node == ztp_node and left_node != ztp_node:
                ztp_interfaces[(left_node.casefold(), left_port.casefold())] = right_port

    with open(lldpq_file, encoding="utf-8") as stream:
        source_links = [
            match.groups() for match in _DOT_EDGE_RE.finditer(stream.read())
        ]
    source_links = _apply_air_topology_policy(
        source_links, air_topology_policy,
    )

    def selected_name(source_name):
        if _air_is_ztp_server(source_name):
            return ztp_node
        value = str(source_name)
        if value.casefold().startswith(_AIR_HOSTNAME_PREFIX.casefold()):
            value = value[len(_AIR_HOSTNAME_PREFIX):]
        return selected.get(value.casefold())

    connected = []
    unconnected = []
    seen_source_endpoints = set()
    for left_source, left_port, right_source, right_port in source_links:
        for source_name, interface in (
            (left_source, left_port), (right_source, right_port)
        ):
            endpoint = (source_name.casefold(), interface.casefold())
            if endpoint in seen_source_endpoints:
                raise ValueError(
                    "LLDPQ interface is connected more than once: "
                    f"{source_name}:{interface}"
                )
            seen_source_endpoints.add(endpoint)

        left_node = selected_name(left_source)
        right_node = selected_name(right_source)
        if left_node and right_node:
            if left_node == ztp_node or right_node == ztp_node:
                if left_node == ztp_node:
                    switch_node, switch_port = right_node, right_port
                    ztp_on_left = True
                else:
                    switch_node, switch_port = left_node, left_port
                    ztp_on_left = False
                ztp_port = ztp_interfaces.get(
                    (switch_node.casefold(), switch_port.casefold())
                )
                if ztp_port is None:
                    unconnected.append((switch_node, switch_port))
                    continue
                if ztp_on_left:
                    connected.append((ztp_node, ztp_port, switch_node, switch_port))
                else:
                    connected.append((switch_node, switch_port, ztp_node, ztp_port))
            else:
                connected.append((left_node, left_port, right_node, right_port))
        elif left_node and left_node != ztp_node:
            unconnected.append((left_node, left_port))
        elif right_node and right_node != ztp_node:
            unconnected.append((right_node, right_port))

    connected.sort(key=_edge_sort_key)
    unconnected.sort(key=lambda item: (_natural_key(item[0]), _natural_key(item[1])))
    return connected, unconnected


def _stable_air_mac(namespace, node, interface):
    """Build a deterministic, locally administered unicast MAC address."""
    canonical_node = str(node)
    if canonical_node.casefold().startswith(_AIR_HOSTNAME_PREFIX.casefold()):
        canonical_node = canonical_node[len(_AIR_HOSTNAME_PREFIX):]
    digest = hashlib.sha256(
        f"{namespace}\0{canonical_node}\0{interface}".encode("utf-8")
    ).digest()
    octets = [0x02, digest[0], digest[1], digest[2], digest[3], digest[4]]
    return ":".join(f"{value:02X}" for value in octets)


def _air_node_layer(name):
    """Return a simple visual layer compatible with the AIR JSON canvas."""
    lowered = name.casefold()
    if "spine" in lowered or "core" in lowered:
        return 0
    if "border" in lowered or "leaf" in lowered or "tor" in lowered:
        return 1
    return 2


def _air_positions(node_names):
    """Lay out nodes deterministically in 275-pixel rows."""
    grouped = {0: [], 1: [], 2: []}
    for name in node_names:
        grouped[_air_node_layer(name)].append(name)
    positions = {}
    for layer, names in grouped.items():
        for index, name in enumerate(names):
            positions[name] = {"x": 825 + index * 275, "y": layer * 275}
    return positions


def _air_oob_config(template_oob, management_ips):
    """Copy OOB defaults and align its /24 subnet with generated management IPs."""
    oob = deepcopy(template_oob)
    if oob is False or oob is None:
        return oob
    if not isinstance(oob, dict):
        raise ValueError(
            "AIR JSON template content.oob must be an object, false, or null"
        )
    addresses = [ipaddress.ip_address(value) for value in management_ips]
    if not addresses:
        return oob
    if not all(isinstance(value, ipaddress.IPv4Address) for value in addresses):
        raise ValueError("AIR JSON currently supports IPv4 management addresses only")
    network = ipaddress.ip_network(f"{addresses[0]}/24", strict=False)
    if any(value not in network for value in addresses):
        raise ValueError("AIR management IP addresses must be in one /24 subnet")
    gateway = network.network_address + 1
    oob["subnets"] = {str(network): {"gateway_ip": str(gateway)}}
    return oob


def _expand_air_port_specs(model, specs):
    """Expand explicit model port specs such as ``swp1-52``."""
    if not isinstance(specs, list):
        raise ValueError(f"AIR template model_port_profiles.{model} must be a list")
    ports = set()
    for raw_spec in specs:
        spec = str(raw_spec).strip()
        if not spec:
            raise ValueError(
                f"AIR template model_port_profiles.{model} contains an empty port"
            )
        match = re.fullmatch(r"([A-Za-z_-]+)(\d+)-(\d+)", spec)
        if match:
            prefix, start_text, end_text = match.groups()
            start, end = int(start_text), int(end_text)
            if start > end or end - start > 4096:
                raise ValueError(
                    f"AIR template model_port_profiles.{model} has invalid range {spec}"
                )
            expanded = {f"{prefix}{number}" for number in range(start, end + 1)}
        else:
            expanded = {spec}
        overlap = ports & expanded
        if overlap:
            raise ValueError(
                f"AIR template model_port_profiles.{model} repeats ports: "
                + ", ".join(sorted(overlap, key=_natural_key))
            )
        ports.update(expanded)
    return ports


def _air_model_port_profiles(template_content):
    """Return complete interface inventories keyed by labels.model.

    AIR export templates encode a model's interface inventory through link
    endpoints, including ``unconnected`` entries. Multiple prototype roles may
    share one model; their non-empty inventories must agree. A newly added role
    (for example OOB-Core) can therefore inherit ports solely through model.
    """
    template_nodes = template_content.get("nodes", {})
    template_links = template_content.get("links", [])
    if not isinstance(template_nodes, dict) or not isinstance(template_links, list):
        return {}

    node_models = {}
    node_ports = {str(name): set() for name in template_nodes}
    for name, node in template_nodes.items():
        labels = node.get("labels") if isinstance(node, dict) else None
        model = labels.get("model") if isinstance(labels, dict) else None
        if str(model or "").strip():
            node_models[str(name)] = str(model).strip()

    for link in template_links:
        if not isinstance(link, list):
            continue
        for endpoint in link:
            if not isinstance(endpoint, dict):
                continue
            node = str(endpoint.get("node") or "")
            interface = str(endpoint.get("interface") or "").strip()
            if node in node_ports and interface:
                node_ports[node].add(interface)

    profiles = {}
    sources = {}
    explicit_profiles = template_content.get("model_port_profiles", {})
    if explicit_profiles is None:
        explicit_profiles = {}
    if not isinstance(explicit_profiles, dict):
        raise ValueError("AIR template content.model_port_profiles must be an object")
    explicit_profile_keys = {
        str(model).strip().casefold() for model in explicit_profiles
        if str(model).strip()
    }
    for model, specs in explicit_profiles.items():
        model_name = str(model).strip()
        if not model_name:
            raise ValueError("AIR template model_port_profiles contains an empty model")
        ports = _expand_air_port_specs(model_name, specs)
        if not ports:
            raise ValueError(
                f"AIR template model_port_profiles.{model_name} has no ports"
            )
        key = model_name.casefold()
        if key in profiles:
            raise ValueError(f"duplicate AIR model port profile: {model_name}")
        profiles[key] = ports
        sources[key] = "content.model_port_profiles"

    for node, model in node_models.items():
        ports = node_ports[node]
        if not ports:
            continue
        key = model.casefold()
        # Explicit model profiles are authoritative; legacy role link entries
        # remain in the export template only as backward-compatible examples.
        if key in explicit_profile_keys:
            continue
        if key in profiles and profiles[key] != ports:
            raise ValueError(
                f"AIR template model {model} has inconsistent port inventories: "
                f"{sources[key]} and {node}"
            )
        profiles[key] = set(ports)
        sources[key] = node
    return profiles


def _air_profile_port_is_used(profile_port, connected_ports):
    """True if a connected interface consumes this model-level base port."""
    profile_key = str(profile_port).casefold()
    connected = {str(port).casefold() for port in connected_ports}
    if profile_key in connected:
        return True
    # A breakout lane such as swp1s0 consumes the model's parent swp1; do not
    # emit a simultaneous ``swp1 unconnected`` entry.
    if re.fullmatch(r"swp\d+", profile_key):
        return any(re.fullmatch(re.escape(profile_key) + r"s\d+", port)
                   for port in connected)
    return False


def generate_air_json(
    air_file,
    json_file,
    template_file=AIR_JSON_TEMPLATE,
    *,
    lldpq_file=None,
    air_topology_policy=None,
):
    """Build AIR JSON from AIR nodes and, when supplied, full LLDPQ links."""
    nodes, air_links = _parse_air_dot(air_file)
    if lldpq_file is None:
        links = air_links
        lldpq_unconnected = []
    else:
        links, lldpq_unconnected = _air_json_links_from_lldpq(
            lldpq_file, nodes, air_links,
            air_topology_policy=air_topology_policy,
        )
    with open(template_file, encoding="utf-8") as stream:
        template = json.load(stream)

    template_content = template.get("content", {})
    template_nodes = template_content.get("nodes", {})
    if not isinstance(template_nodes, dict) or not template_nodes:
        raise ValueError(f"AIR JSON template has no node defaults: {template_file}")
    template_oob = template_content.get("oob", {})
    model_port_profiles = _air_model_port_profiles(template_content)
    explicit_model_profile_keys = {
        str(model).strip().casefold()
        for model in (template_content.get("model_port_profiles") or {})
        if str(model).strip()
    }
    positions = _air_positions([name for name, _attrs in nodes])

    json_nodes = {}
    management_ips = []
    used_management_ips = set()
    used_macs = set()
    for name, attrs in nodes:
        try:
            cpu = int(attrs.get("cpus", "1"))
            memory = int(attrs.get("memory", "1024"))
        except ValueError as error:
            raise ValueError(f"invalid cpu or memory for AIR node {name}") from error
        management_ip = attrs.get("mgmt_ip", "").strip()
        if management_ip:
            ipaddress.ip_address(management_ip)
            if management_ip in used_management_ips:
                raise ValueError(f"duplicate AIR management IP: {management_ip}")
            used_management_ips.add(management_ip)
            management_ips.append(management_ip)

        _template_name, node = _air_template_node(
            template_nodes,
            name,
            explicit_name=attrs.get("template_node"),
        )
        inherit_oob_leaf = "oobofoob" in name.casefold()
        if not inherit_oob_leaf:
            node["cpu"] = cpu
            node["memory"] = memory
        node["positioning"] = positions[name]
        # OOBofOOB inherits the OOB leaf's VM shape and model inventory, not
        # its template image version.  The DOT value carries the explicit
        # project --os-version selected by load and must win for every
        # generated Cumulus node.
        node["os"] = attrs.get("os", node.get("os", "")).strip()
        # The ZTP server template deliberately carries a fixed eth0 MAC so
        # AIR can expose its outbound management connection predictably.
        # Other nodes keep their deterministic per-host management MAC.
        template_management_interfaces = deepcopy(
            node.get("management_interfaces") or {}
        )
        template_eth0 = template_management_interfaces.get("eth0") or {}
        template_eth0_mac = str(
            template_eth0.get("mac_address") or ""
        ).strip()
        if _air_is_ztp_server(name) and template_eth0_mac:
            management_mac = template_eth0_mac
        else:
            management_mac = _stable_air_mac("management", name, "eth0")
        if management_mac in used_macs:
            raise ValueError(f"duplicate generated AIR MAC: {management_mac}")
        used_macs.add(management_mac)
        management_interfaces = template_management_interfaces
        management_interfaces["eth0"] = {
            "ip": management_ip or None,
            "mac_address": management_mac,
        }
        node["management_interfaces"] = management_interfaces
        json_nodes[name] = node

    def link_mac(node_name, interface):
        """Return an endpoint MAC and whether it is shared with management."""
        if interface.casefold() == "eth0":
            return (
                json_nodes[node_name]["management_interfaces"]["eth0"][
                    "mac_address"
                ],
                True,
            )
        return _stable_air_mac("link", node_name, interface), False

    json_links = []
    used_endpoints = set()
    connected_by_node = {name.casefold(): set() for name in json_nodes}
    for left_node, left_port, right_node, right_port in links:
        link = []
        for node, interface in (
            (left_node, left_port), (right_node, right_port)
        ):
            endpoint = (node.casefold(), interface.casefold())
            if endpoint in used_endpoints:
                raise ValueError(
                    f"AIR interface is connected more than once: {node}:{interface}"
                )
            used_endpoints.add(endpoint)
            connected_by_node.setdefault(node.casefold(), set()).add(interface)
            mac, reuses_management_mac = link_mac(node, interface)
            if not reuses_management_mac and mac in used_macs:
                raise ValueError(f"duplicate generated AIR MAC: {mac}")
            if not reuses_management_mac:
                used_macs.add(mac)
            link.append({"interface": interface, "node": node, "mac": mac})
        json_links.append(link)

    # Every retained LLDPQ endpoint whose peer node is outside the AIR node
    # allowlist remains visible as an exact unconnected interface.
    unconnected_count = 0
    for name, interface in lldpq_unconnected:
        endpoint = (name.casefold(), interface.casefold())
        if endpoint in used_endpoints:
            continue
        used_endpoints.add(endpoint)
        connected_by_node.setdefault(name.casefold(), set()).add(interface)
        mac, reuses_management_mac = link_mac(name, interface)
        if not reuses_management_mac and mac in used_macs:
            raise ValueError(f"duplicate generated AIR MAC: {mac}")
        if not reuses_management_mac:
            used_macs.add(mac)
        json_links.append([
            {"interface": interface, "node": name, "mac": mac},
            "unconnected",
        ])
        unconnected_count += 1

    # Server nodes are intentionally removed from AIR, but generate_air_dot()
    # records their exact switch-side ports on the node.  Materialize those
    # endpoints as unconnected before completing the remaining model profile;
    # adding them to connected_by_node also prevents a breakout parent (swp11)
    # from being emitted alongside its preserved lanes (swp11s0..swp11s3).
    for name, attrs in nodes:
        raw_ports = attrs.get("preserved_unconnected_ports", "").strip()
        if not raw_ports:
            continue
        ports = [port.strip() for port in raw_ports.split(",") if port.strip()]
        if len({port.casefold() for port in ports}) != len(ports):
            raise ValueError(
                f"AIR node {name} repeats preserved unconnected ports"
            )
        for interface in ports:
            endpoint = (name.casefold(), interface.casefold())
            if endpoint in used_endpoints:
                # LLDPQ-derived endpoints supersede the compatibility
                # preserved_unconnected_ports attribute in AIR DOT.
                if lldpq_file is not None:
                    continue
                raise ValueError(
                    f"AIR preserved interface is already connected: "
                    f"{name}:{interface}"
                )
            used_endpoints.add(endpoint)
            connected_by_node.setdefault(name.casefold(), set()).add(interface)
            mac, reuses_management_mac = link_mac(name, interface)
            if not reuses_management_mac and mac in used_macs:
                raise ValueError(f"duplicate generated AIR MAC: {mac}")
            if not reuses_management_mac:
                used_macs.add(mac)
            json_links.append([
                {"interface": interface, "node": name, "mac": mac},
                "unconnected",
            ])
            unconnected_count += 1

    # Preserve explicit outbound interfaces from the AIR template. They are
    # not physical P2P links and therefore cannot be reconstructed from DOT.
    # Only endpoints whose node is present in the generated topology apply.
    template_links = template_content.get("links", [])
    if not isinstance(template_links, list):
        raise ValueError("AIR template content.links must be a list")
    for template_link in template_links:
        if not (
            isinstance(template_link, list)
            and len(template_link) == 2
            and template_link[1] == "outbound"
            and isinstance(template_link[0], dict)
        ):
            continue
        source = template_link[0]
        node_name = str(source.get("node") or "").strip()
        interface = str(source.get("interface") or "").strip()
        mac = str(source.get("mac") or "").strip()
        if not node_name or not interface or not mac:
            raise ValueError(
                "AIR template outbound endpoint requires node/interface/mac"
            )
        if node_name not in json_nodes:
            continue
        endpoint = (node_name.casefold(), interface.casefold())
        if endpoint in used_endpoints:
            raise ValueError(
                f"AIR template outbound interface is already connected: "
                f"{node_name}:{interface}"
            )
        if interface.casefold() == "eth0":
            mac, reuses_management_mac = link_mac(node_name, interface)
        else:
            reuses_management_mac = False
        if not reuses_management_mac and mac in used_macs:
            raise ValueError(f"duplicate AIR outbound MAC: {mac}")
        used_endpoints.add(endpoint)
        if not reuses_management_mac:
            used_macs.add(mac)
        connected_by_node.setdefault(node_name.casefold(), set()).add(interface)
        json_links.append([
            {"interface": interface, "node": node_name, "mac": mac},
            "outbound",
        ])

    # Complete each generated node from its hardware model inventory. Template
    # role names never appear here: OOB-Core and TAN-Spine both obtain SN5610's
    # port list through labels.model. Templates without model inventories retain
    # the legacy actual-links-only behavior.
    for name, node in json_nodes.items():
        labels = node.get("labels") if isinstance(node, dict) else None
        model = labels.get("model") if isinstance(labels, dict) else None
        profile = model_port_profiles.get(str(model or "").casefold())
        if not profile:
            continue
        connected_ports = connected_by_node.get(name.casefold(), set())
        # Only explicitly declared inventories are strict contracts. Inferred
        # legacy profiles may contain AIR aliases (for example 1/2/mgmt on an
        # FW) which cannot safely be normalized without model-specific rules.
        if str(model or "").casefold() in explicit_model_profile_keys:
            invalid_connected = [
                interface for interface in connected_ports
                if not any(_air_profile_port_is_used(port, {interface})
                           for port in profile)
            ]
            if invalid_connected:
                raise ValueError(
                    f"AIR node {name} uses ports outside model {model} profile: "
                    + ", ".join(sorted(invalid_connected, key=_natural_key))
                )
        for interface in sorted(profile, key=_natural_key):
            if _air_profile_port_is_used(interface, connected_ports):
                continue
            endpoint = (name.casefold(), interface.casefold())
            if endpoint in used_endpoints:
                continue
            used_endpoints.add(endpoint)
            mac, reuses_management_mac = link_mac(name, interface)
            if not reuses_management_mac and mac in used_macs:
                raise ValueError(f"duplicate generated AIR MAC: {mac}")
            if not reuses_management_mac:
                used_macs.add(mac)
            json_links.append([
                {"interface": interface, "node": name, "mac": mac},
                "unconnected",
            ])
            unconnected_count += 1

    document = {
        "format": template.get("format", "JSON"),
        "ztp": template.get("ztp"),
        "content": {
            "nodes": json_nodes,
            "links": json_links,
            "oob": _air_oob_config(template_oob, management_ips),
        },
        "name": os.path.basename(json_file)[:-5],
    }

    os.makedirs(os.path.dirname(os.path.abspath(json_file)), exist_ok=True)
    temporary = f"{json_file}.tmp.{os.getpid()}"
    try:
        with open(temporary, "w", encoding="utf-8") as stream:
            json.dump(document, stream, ensure_ascii=False, indent=4)
            stream.write("\n")
        os.chmod(temporary, 0o644)
        os.replace(temporary, json_file)
    finally:
        if os.path.exists(temporary):
            os.unlink(temporary)

    print(
        f"Generated: {json_file}  ({len(json_nodes)} nodes, "
        f"{len(links)} connected links, {unconnected_count} unconnected ports)"
    )
    return json_file


if __name__ == "__main__":
    main()
